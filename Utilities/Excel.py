from openpyxl import load_workbook
from openpyxl.workbook import Workbook


class Excel:
    def __init__(self, file_path, sheet_name):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.workbook = None

    def create(self):
        self.workbook = Workbook()
        self.workbook.save(self.file_path)

    def open(self):
        print("Opening workbook " + self.file_path)
        try:
            self.workbook = load_workbook(self.file_path)
        except FileNotFoundError:
            print("File not found. Creating a new workbook.")


    def save(self):
        self.workbook.save(self.file_path)

    # change single Cell, ex: (A10) A = Columb Row: 10
    def change_cell_content(self, column_letter,  row_number, new_content):
        cell = str(column_letter) + str(row_number)

        # Get the sheet by name
        sheet = self.workbook[self.sheet_name]

        # Update the value of the specified cell
        sheet[cell] = new_content

